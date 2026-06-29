#!/usr/bin/env python3
"""
Step 3: Parse environmental sensor data from sensor_data.xlsx files.

Structure:
  - Each sheet = one date (e.g., '2024-09-05')
  - Each row = 30-min timestamp
  - Columns = sensor_type(process_group), e.g. 'Air Temperature(Process1)'
  - Multiple process groups in one file

We compute daily mean across ALL process groups for each sensor type.
"""
import os, glob, csv, re
import openpyxl
from collections import defaultdict
from datetime import datetime

BASE = r'/Users/rainxu/Downloads/2023-2025 Tomato dataset'

SENSOR_KEYWORDS = {
    'Air_Temperature': ['Air Temperature', '空气温度'],
    'Relative_Humidity': ['Relative Humidity', '相对湿度'],
    'Light_Intensity': ['Light Intensity', '光照'],
    'CO2': ['CO2', '二氧化碳', 'carbon dioxide'],
    'Soil_Moisture': ['Soil Moisture', '土壤水分', '土壤湿度'],
    'Soil_Temperature': ['Soil Temperature', '土壤温度'],
}


def classify_column(col_name):
    """Return sensor type key if column matches, or None."""
    col_lower = col_name.lower()
    for key, keywords in SENSOR_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in col_lower:
                return key
    return None


def parse_and_dump_sheets(filepath):
    """Parse all sheets (dates) in a sensor xlsx."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    daily_values = defaultdict(lambda: defaultdict(list))

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Sheet name might be the date itself (e.g., '2024-09-05')
        date_key = None
        try:
            dt = datetime.strptime(sheet_name, '%Y-%m-%d')
            date_key = dt.strftime('%Y%m%d')
        except ValueError:
            pass
        # Or '2024-09-05' → '20240905'
        try:
            dt = datetime.strptime(sheet_name.replace('-', ''), '%Y%m%d')
            date_key = dt.strftime('%Y%m%d')
        except:
            pass

        if date_key is None:
            continue

        # Build column classification from header
        header_row = None
        col_types = {}
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri == 0:
                for ci, cell in enumerate(row):
                    col_name = str(cell).strip() if cell else ''
                    sensor_type = classify_column(col_name)
                    if sensor_type:
                        col_types[ci] = sensor_type
                header_row = ri
                break

        if not col_types:
            continue

        # Process data rows
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri <= (header_row or 0):
                continue
            if not row or not row[0]:
                continue

            for ci, val in enumerate(row):
                if ci in col_types and val is not None:
                    try:
                        fval = float(val)
                        if not (fval != fval):  # not NaN
                            daily_values[date_key][col_types[ci]].append(fval)
                    except (ValueError, TypeError):
                        pass

    # Aggregate to daily means
    result = {}
    for date_key, fields in daily_values.items():
        result[date_key] = {'Date': date_key}
        for field in ['Air_Temperature', 'Relative_Humidity', 'Light_Intensity',
                       'CO2', 'Soil_Moisture', 'Soil_Temperature']:
            vals = fields.get(field, [])
            result[date_key][field] = round(sum(vals) / len(vals), 4) if vals else ''

    return result


def main():
    out_path = os.path.join(os.path.dirname(__file__), 'sensor_daily.csv')

    all_daily = {}

    for year in ['2023', '2024', '2025']:
        if year == '2023':
            pattern = os.path.join(BASE, year, '*', 'processed_data', 'sensor_data', 'sensor_data.xlsx')
        else:
            sub = '2024' if year == '2024' else '2025'
            pattern = os.path.join(BASE, year, sub, '*', 'processed_data', 'sensor_data', 'sensor_data.xlsx')

        files = glob.glob(pattern)
        print(f'{year}: {len(files)} sensor file(s)')

        for fpath in files:
            data = parse_and_dump_sheets(fpath)
            for date_key, fields in data.items():
                if date_key not in all_daily:
                    fields['Source'] = year
                    all_daily[date_key] = fields
                else:
                    # Merge: keep existing, update with new non-empty values
                    for fld, val in fields.items():
                        if val != '' and val is not None:
                            all_daily[date_key][fld] = val

    if not all_daily:
        print('ERROR: No dates parsed. Check sensor file structure.')
        print(f'Files found: {files}')
        # Debug: print first file's structure
        if files:
            from openpyxl import load_workbook
            wb = load_workbook(files[0], read_only=True)
            print(f'First file sheets: {wb.sheetnames[:5]}')
            ws = wb[wb.sheetnames[0]]
            print(f'First sheet header (col 0-5): {[str(c)[:40] for c in list(ws.iter_rows(values_only=True))[0]][:5]}')
        return

    # Write CSV
    fieldnames = ['Date', 'Source', 'Air_Temperature', 'Relative_Humidity',
                  'Light_Intensity', 'CO2', 'Soil_Moisture', 'Soil_Temperature']
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for date_key in sorted(all_daily.keys()):
            w.writerow(all_daily[date_key])

    print(f'\nParsed {len(all_daily)} unique dates')
    print(f'Date range: {sorted(all_daily.keys())[0]} to {sorted(all_daily.keys())[-1]}')
    print(f'Sample: {dict(list(all_daily.items())[:3])}')
    print(f'Written to {out_path}')


if __name__ == '__main__':
    main()
